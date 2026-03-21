# exporter.py
from __future__ import annotations

import os
from typing import List, Any, Tuple
from PySide6.QtWidgets import QTreeView
from PySide6.QtCore import QModelIndex

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _collect_rows_with_depth(view: QTreeView) -> List[Tuple[List[Any], int]]:
    """ 각 행의 데이터와 해당 행의 계층 깊이(depth)를 함께 수집합니다. """
    model = view.model()
    if model is None:
        return []

    rows: List[Tuple[List[Any], int]] = []

    def walk(parent: QModelIndex, depth: int):
        for r in range(model.rowCount(parent)):
            idx0 = model.index(r, 0, parent)
            row_vals = []
            for c in range(model.columnCount(parent)):
                idx = model.index(r, c, parent)
                row_vals.append(model.data(idx))
            rows.append((row_vals, depth))
            walk(idx0, depth + 1)

    walk(QModelIndex(), 0)
    return rows


def export_current_view_to_excel(view: QTreeView, path: str) -> None:
    """ 전문 보고서 수준의 스타일링과 숫자 형식이 적용된 엑셀 파일을 생성합니다. """
    rows_data = _collect_rows_with_depth(view)
    if not rows_data:
        raise ValueError("내보낼 데이터가 없습니다.")

    model = view.model()
    headers = [model.headerData(i, view.header().orientation(), role=0) for i in range(model.columnCount())]

    wb = Workbook()
    ws = wb.active
    ws.title = "예산분석결과"

    # --- 스타일 정의 ---
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(bold=True)
    
    # 계층별 배경색 (Depth 0: 짙은 회색, Depth 1: 연한 회색, Depth 2+: 흰색)
    fills = [
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # Depth 0
        PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid"),  # Depth 1
    ]
    
    border = Border(
        left=Side(style='thin', color='B2B2B2'),
        right=Side(style='thin', color='B2B2B2'),
        top=Side(style='thin', color='B2B2B2'),
        bottom=Side(style='thin', color='B2B2B2')
    )

    # --- 헤더 작성 ---
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # --- 데이터 작성 ---
    for row_idx, (vals, depth) in enumerate(rows_data, 2):
        for col_idx, val in enumerate(vals, 1):
            # 1. 수치 데이터 변환 (콤마 제거 후 숫자형으로)
            clean_val = val
            is_number = False
            if isinstance(val, str) and col_idx > 1: # 항목명 제외한 컬럼들
                try:
                    # 콤마 제거 및 소수점 처리
                    temp_val = val.replace(',', '').replace(' ', '')
                    if temp_val:
                        clean_val = float(temp_val)
                        is_number = True
                except ValueError:
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=clean_val)
            
            # 2. 계층별 스타일링
            if depth < len(fills):
                cell.fill = fills[depth]
            if depth < 2:
                cell.font = Font(bold=True)
            
            # 3. 정렬 및 서식
            if is_number:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                # 항목명 들여쓰기 시각화 (엑셀 내장 들여쓰기 속성 활용)
                if col_idx == 1:
                    cell.alignment = Alignment(indent=depth * 2, vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell.border = border

    # --- 후처리: 너비 조절 및 틀 고정 ---
    # 각 컬럼의 최대 길이를 계산하여 너비 자동 조절
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        adjusted_width = (max_length + 4) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 60) # 최대 60

    ws.freeze_panes = "A2" # 첫 행 고정
    
    # 저장
    wb.save(path)
